import datetime

import requests
from bs4 import BeautifulSoup
import os
import openpyxl

url = "https://kramp.ru/ajax/?controller=user&action=auth"

with open('auth.txt', 'r', encoding='utf-8') as file:
    auth_data = [i.replace('\n', '') for i in file]

login = auth_data[0]
password = auth_data[1]

payload = {
    'USER_LOGIN': login,
    'USER_PASSWORD': password,
    'USER_REMEMBER': 'Y',
    'backurl': '/',
    'AUTH_FORM': 'Y',
    'PROFILE_URL': '/',
    'TYPE': 'AUTH',
}

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36',
}

session = requests.session()
response1 = session.post(url=url, data=payload, headers=headers)

exel_name = [each for each in os.listdir(os.getcwd()) if each.endswith('.xlsx')][0]
book = openpyxl.load_workbook(exel_name)
sheet = book.active
row = 1
articles = []
file_name = exel_name.split('.')[0]
while True:
    article = sheet[row][0].value
    row += 1
    if article is None:
        break
    articles.append(str(article))

book.close()
items_dict = {}
now = str(datetime.datetime.now().strftime('%d.%m.%Y'))[:10]

for article in articles:
    print(f'Поиск артикула {article}..')
    response = session.get(f'https://kramp.ru/search/{article}/', headers=headers)
    soup = BeautifulSoup(response.text, 'lxml')
    try:
        item = soup.find('div', class_='catalog-item')
        item_link = 'https://kramp.ru' + item.find('a')['href']
        item_title = str(item.find('div', class_='text-group').find(class_='item-title').text)
    except:
        items_dict[article] = ['-', '-', now, '-']
        continue
    if item_title.upper() == article.upper():
        response = session.get(item_link, headers=headers)
        soup = BeautifulSoup(response.text, 'lxml')
        try:
            photo = soup.find(class_='block__pic')['src']
        except:
            photo = '-'
        info_block = soup.find('div', 'card-price-block')
        price_text = info_block.find(class_='card-price-block__price').text.split('р')[0].replace('', '')
        price = ''.join([i for i in price_text if i.isdigit() or i=='.'])
        availability_text = info_block.find(class_='product-info__row').text
        availability = ''.join([i for i in availability_text if i.isdigit()])
        item_name = soup.find('h1').text.split('артикул')[0].replace('\n', '').strip()
        items_dict[article] = [item_name, price, now, availability, photo]

print('[+][+][+][+]ЗАПИСЬ СОБРАННЫХ ДАННЫХ В ФАЙЛ[+][+][+][+]')
items_count = len(items_dict)
book = openpyxl.Workbook()
sheet1 = book.active

row = 1
count = 0
for article in items_dict:
    sheet1.cell(row, 1).value = article
    sheet1.cell(row, 2).value = article + " " + items_dict[article][0]
    sheet1.cell(row, 3).value = items_dict[article][1]
    sheet1.cell(row, 4).value = items_dict[article][2]
    sheet1.cell(row, 5).value = items_dict[article][3]
    sheet1.cell(row, 6).value = items_dict[article][4]
    row += 1
    count += 1
    print(f'Осталось записать значений: {items_count - count}')
    continue

book.save(f'result_{file_name}.xlsx')
book.close()
