import datetime
import time

from selenium import webdriver
from selenium.webdriver.common.by import By
import os
import openpyxl

url = 'https://kramp.ru/'

exel_name = [each for each in os.listdir(os.getcwd()) if each.endswith('.xlsx')][0]
book = openpyxl.load_workbook(exel_name)
sheet = book.active
row = 1
articles = []
file_name = exel_name.split('.')[0]
while True:
    article = sheet[row][0].value
    row += 1
    if article is None:
        break
    articles.append(str(article))

book.close()

options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")

with webdriver.Chrome(options=options) as browser:
    browser.get(url)
    try:
        modal_window = browser.find_elements(By.CSS_SELECTOR, 'div.modal-content')[2]
        modal_window.find_element(By.TAG_NAME, 'a').click()
    except:
        pass
    auth_button = browser.find_element(By.CLASS_NAME, 'top-auth').find_element(By.TAG_NAME, 'a')
    auth_button.click()
    time.sleep(2)
    login_input = browser.find_element(By.ID, 'login')
    password_input = browser.find_element(By.ID, 'password')
    button_login = browser.find_elements(By.CSS_SELECTOR, 'button.submit')[1]
    login_input.send_keys('2@seltop.ru')
    password_input.send_keys('seltop77')
    time.sleep(1)
    button_login.click()
    time.sleep(2)
    items_dict = {}
    now = str(datetime.datetime.now())[:10]
    for article in articles:
        input_field = browser.find_element(By.ID, 'search-control')
        input_field.clear()
        input_field.click()
        input_field.send_keys(article)
        search_but = browser.find_element(By.CSS_SELECTOR, 'button.btn-info')
        search_but.click()
        try:
            item = browser.find_element(By.CSS_SELECTOR, 'div.catalog-item')
            item_link = item.find_element(By.TAG_NAME, 'a').get_attribute('href')
            item_title = str(item.find_element(By.CSS_SELECTOR, 'div.text-group').find_element(By.CLASS_NAME,
                                                                                           'item-title').text)
        except:
            items_dict[article] = ['-', '-', now, '-']
            continue
        if item_title.upper() == article.upper():
            browser.get(item_link)
            try:
                photo = browser.find_element(By.CLASS_NAME, 'block__pic').get_attribute('src')
            except:
                photo = '-'
            info_block = browser.find_element(By.CSS_SELECTOR, 'div.col-sm-5')
            price = info_block.find_element(By.CLASS_NAME, 'card-price-block__price').text.split('р')[0].strip()
            availability_text = info_block.find_element(By.CLASS_NAME, 'product-info__row').text
            availability = ''.join([i for i in availability_text if i.isdigit()])
            item_name = browser.find_element(By.TAG_NAME, 'h1').text.split('артикул')[0]
            items_dict[article] = [item_name, price, now, availability, photo]
        else:
            continue

print('[+][+][+][+]ЗАПИСЬ СОБРАННЫХ ДАННЫХ В ФАЙЛ[+][+][+][+]')
items_count = len(items_dict)
book = openpyxl.Workbook()
sheet1 = book.active

row = 1
count = 0
for article in items_dict:
    sheet1.cell(row, 1).value = article
    sheet1.cell(row, 2).value = article + " " + items_dict[article][0]
    sheet1.cell(row, 3).value = items_dict[article][1]
    sheet1.cell(row, 4).value = items_dict[article][2]
    sheet1.cell(row, 5).value = items_dict[article][3]
    sheet1.cell(row, 6).value = items_dict[article][4]
    row += 1
    count += 1
    print(f'Осталось записать значений: {items_count - count}')
    continue

book.save(f'result_{file_name}.xlsx')
book.close()
